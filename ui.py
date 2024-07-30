from window import Ui_MainWindow
from rws import RWS, Vibration, GohomeThread,GovibrationposThread
from mti import Receiver, parser
from dataprocess import calculation,initt
import sys,os
from openpyxl import load_workbook
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox


class UI(QMainWindow, Ui_MainWindow):

    def __init__(self):
        params = initt()
        self.path = params['path']
        # self.validFormat = '<font color="green" size="12">{}</font>'
        super(UI, self).__init__()
        self.setupUi(self)
        self.rws = RWS(url=params['url'])
        self.viration_test_button.clicked.connect(self.START)
        self.get_serial_number_button.clicked.connect(self.get_serial)
        self.stop_button.clicked.connect(self.stop)
        self.gotosyncpose_button.clicked.connect(self.gohome)
        self.motor_on_button.clicked.connect(self.motor_on)

    def run(self):
        self.vibration = Vibration(self.rws)
        self.vibration.update_status.connect(self.update_status)
        self.vibration.stop_record.connect(self.mti_receive.stop_cycle)
        self.vibration.error.connect(self.get_controller_error)
        self.vibration.start()

    def run_sensor(self):
        # if self.validateInput():
        #     return
        # self.clear_status()
        # self.go_vibration_pos()

        self.mti_receive = Receiver(series=self.serial_number.text(),path=self.path)
        self.mti_receive.update_status.connect(self.update_status)
        self.mti_receive.start_controller.connect(self.run)
        self.mti_receive.error.connect(self.get_sensor_error)
        self.mti_receive.parser.connect(self.run_parser_and_dataprocess)
        self.mti_receive.start()

    def START(self):
        if self.validateInput():
            return
        self.clear_status()
        self.govibrationpos = GovibrationposThread(self.rws)
        self.govibrationpos.update_status.connect(self.update_status)
        self.govibrationpos.error.connect(self.go_vibration_pose_rapid)
        self.govibrationpos.start_record.connect(self.run_sensor)
        self.govibrationpos.start()

    def run_parser_and_dataprocess(self):
        try:
            self.update_status('开始解析数据')
            parser(filename=self.serial_number.text(),path=self.path)
            self.update_status('数据解析完成')
            self.update_status('开始结果分析')
            pass_or_not = calculation(self.serial_number.text(),self.path)#'1100-502179'
            self.update_status('结果分析完成')
            print(pass_or_not)
            # self.update_status(str(pass_or_not))
            pass_or_not.insert(0,self.serial_number.text())
            self.write_data(pass_or_not)
            print('PASS')
            self.run_complete()
        except Exception as e:
            self.update_status('运行是发生错误,如下：')
            self.update_status(repr(e))
    
    def run_complete(self):
        self.status_text.append('Vibration测试结束！')

    def update_status(self, message):
        self.status_text.append(message)

    def get_controller_error(self, value):
        choice = QMessageBox.warning(
            self, "警告", value, QMessageBox.Yes | QMessageBox.No
        )
        if choice == QMessageBox.Yes:
            self.vibration.stop_cycle()
        else:
            self.vibration.exit()

    def get_sensor_error(self, value):
        choice = QMessageBox.warning(
            self, "警告", value, QMessageBox.Yes | QMessageBox.No
        )
        if choice == QMessageBox.Yes:
            self.mti_receive.stop_cycle()
        else:
            self.mti_receive.exit()
    
    def go_home_rapid(self, value):
        choice = QMessageBox.warning(
            self, "警告", value, QMessageBox.Yes | QMessageBox.No
        )
        if choice == QMessageBox.Yes:
            self.gohome_run.stop_cycle()
        else:
            self.gohome_run.exit()
    
    def go_vibration_pose_rapid(self, value):
        choice = QMessageBox.warning(
            self, "警告", value, QMessageBox.Yes | QMessageBox.No
        )
        if choice == QMessageBox.Yes:
            self.govibrationpos.stop_cycle()
        else:
            self.govibrationpos.exit()

    def clear_status(self):
        self.status_text.setText('')

    def get_serial(self):
        if self.rws.baseurl == "https://192.168.125.1":
            serial = self.rws.GETserial()
        else:
            serial = "1100-000001"
        self.serial_number.setText(serial)

    def gohome(self):
        self.clear_status()
        self.gohome_run = GohomeThread(self.rws)
        self.gohome_run.update_status.connect(self.update_status)
        self.gohome_run.error.connect(self.go_home_rapid)
        self.gohome_run.start()

    def motor_on(self):
        self.rws.motor("motoron")
        self.update_status("motor on!")

    def stop(self):
        self.rws.stopexcuseRapid()
        # self.update_status("Stop rapid!")
        if hasattr(self,'mti_receive'):
            self.mti_receive.quit()
        if hasattr(self,'vibration'):
            self.vibration.quit()
        self.update_status('程序停止了！')

    def validateInput(self):
        # 获取输入框的文本内容
        text = self.serial_number.text()
        if text == "":
            QMessageBox.warning(self, "Warning", "机器人序列号不能为空")
            return True
        self.series = text
        return False
    
    def judge(self, values):
        i = 1
        tune_dic = {"Kv_2": 0, "Kv_3": 0, "Kp_2": 0, "Kp_3": 0}

        if values[1] == 1 or values[2] == 1:
            i += 1
        if values[3] == 1 or values[4] == 1:
            i += 1
        if values[5] == 1 or values[6] == 1:
            i += 1

        if i == 1:
            pass
        elif i == 2:
            tune_dic["Kv_2"] = 50
            tune_dic["Kv_3"] = 50
        elif i == 3:
            tune_dic["Kp_2"] = 60
            tune_dic["Kp_3"] = 60
            tune_dic["Kv_2"] = 20
            tune_dic["Kv_3"] = 20
        else:
            return False
        
        if tune_dic:
                if tune_dic['Kv_2'] == 0:
                    self.statusWindow.append(self.validFormat.format(f"所有过程均不振动,无需调整参数！"))
                elif tune_dic['Kv_2'] != 0 and tune_dic['Kp_2'] == 0:
                    self.statusWindow.append(self.validFormat.format(f"第一段振动，第二段不振动！需要调整参数如下:"))
                    self.statusWindow.append(self.validFormat.format(f"二轴Kv为{tune_dic['Kv_2']}, 三轴Kv为{tune_dic['Kv_3']}"))
                elif tune_dic['Kv_2'] != 0 and tune_dic['Kp_2'] != 0:
                    self.statusWindow.append(self.validFormat.format(f"第一、二段振动，第三段不振动！需要调整参数如下:"))
                    self.statusWindow.append(self.validFormat.format(f"二轴Kv为{tune_dic['Kv_2']}, 三轴Kv为{tune_dic['Kv_3']}, 二轴Kp为{tune_dic['Kp_2']}, 三轴Kp为{tune_dic['Kp_3']}"))
        else:
            self.statusWindow.append(self.validFormat.format(f"所有阶段都振动，请检查设备！"))

    def write_data(self, values):
        # values.insert(0, self.series)
        result_path = os.path.join(self.path,"RESULT", "RESULT.xlsx")
        wb = load_workbook(result_path)
        sheet = wb.active
        last_row = sheet.max_row
        for index, cell in enumerate(values):
            sheet.cell(row=last_row + 1, column=index + 1).value = cell
        wb.save(result_path)
    
    

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = UI()
    window.show()
    sys.exit(app.exec_())
